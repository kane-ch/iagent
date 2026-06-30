#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
上市公司分部业务财务报表Excel生成工具

支持多层级分部展示（如BABA的三层业务结构），体现层级关系。
包含收入、EBITA等指标，支持YoY同比高亮显示。

用法：
    python generate_segment_excel.py <ticker> [--output <path>]
    python generate_segment_excel.py --json <json_file> [--output <path>]

示例：
    python generate_segment_excel.py BABA
    python generate_segment_excel.py BABA --output ./baba_segments.xlsx
"""

import json
import subprocess
import sys
import os
import argparse
import logging
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─────────────────────────────────────────────────────────────
# 颜色配置 - 按层级区分
# ─────────────────────────────────────────────────────────────

# 一级业务分部的背景色组合（用于区分不同的一级业务）
LEVEL1_COLORS = [
    {'bg': 'D9E1F2', 'font': '2F5496'},    # 浅蓝-深蓝字体
    {'bg': 'E2EFDA', 'font': '548235'},    # 浅绿-深绿字体
    {'bg': 'FFF2CC', 'font': 'BF8F00'},    # 浅黄-深黄字体
    {'bg': 'FCE4D6', 'font': 'C65911'},    # 浅橙-深橙字体
    {'bg': 'E2DAF0', 'font': '7030A0'},    # 浅紫-深紫字体
    {'bg': 'F8CBAD', 'font': 'B45F06'},    # 橙黄-深橙字体
]

COLORS = {
    'header': '4472C4',       # 表头-深蓝
    'header_font': 'FFFFFF',  # 表头字体-白色
    'revenue': 'DEEBF7',      # 收入-浅蓝
    'ebita': 'E2EFDA',        # EBITA-浅绿
    'yoy_decline': 'FF0000',  # YoY下降-红色
    'yoy_growth': '00B050',   # YoY增长-绿色
    'ratio': 'FFF2CC',        # 比率-浅黄
    'default': None,
}

# YoY阈值配置
YOY_DECLINE_THRESHOLD = -5.0   # 下降超过此阈值红色高亮
YOY_GROWTH_THRESHOLD = 30.0    # 增长超过此阈值绿色高亮

# ─────────────────────────────────────────────────────────────
# 辅助函数
# ─────────────────────────────────────────────────────────────

def setup_logging(log_file):
    """设置日志系统"""
    if not os.path.exists(os.path.dirname(log_file)):
        os.makedirs(os.path.dirname(log_file))

    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s',
        filename=log_file,
        filemode='w',
        encoding='utf-8'
    )
    return logging.getLogger(__name__)


def log_stdout(msg):
    """输出到 stdout（极简）"""
    print(msg, flush=True)


def get_segment_display_name(segment, level=None):
    """生成分部显示名称，包含层级缩进"""
    name = segment.get('segmentName', segment.get('segmentId', 'Unknown'))
    level = level or segment.get('level', 1)

    # 根据层级添加缩进
    indent = '  ' * (level - 1)
    prefix = '├─ ' if level > 1 else ''
    return f"{indent}{prefix}{name}"


def get_all_periods(segments):
    """从所有分部中提取所有周期"""
    periods = set()

    def collect_periods(seg):
        for metric in seg.get('metrics', []):
            period = metric.get('period')
            if period:
                periods.add(period)
        for child in seg.get('children', []):
            collect_periods(child)

    for seg in segments:
        collect_periods(seg)

    # 按时间排序（如 2024Q1, 2024Q2, 2024Q3, 2024Q4）
    return sorted(list(periods))


def flatten_segments_by_depth(segments, result=None, parent_level=0):
    """深度优先遍历，将分层的分部列表展平，保持层级关系"""
    if result is None:
        result = []

    for seg in segments:
        level = seg.get('level', parent_level + 1)
        seg['level'] = level
        result.append(seg)

        # 递归处理子分部（深度优先）
        children = seg.get('children', [])
        if children:
            flatten_segments_by_depth(children, result, level)

    return result


def has_any_data(values):
    """检查是否有任何有效数据"""
    return any(v is not None for v in values)


def build_data_rows(segments, periods, logger=None):
    """构建数据行
    展示逻辑：
    1. 按层级深度优先遍历（L1 → L21 → L31 → L32 → L22 → L33）
    2. 每个分部分组展示：收入 → 收入YoY → EBITA → EBITA YoY → EBITA利润率
    3. 空数据行过滤：指标值全为空则不展示
    """
    rows = []

    # 深度优先遍历展平分部，保持层级
    flat_segments = flatten_segments_by_depth(segments)

    if logger:
        logger.info(f"深度优先展平后共 {len(flat_segments)} 个分部，{len(periods)} 个周期")

    for seg in flat_segments:
        level = seg.get('level', 1)
        seg_name = seg.get('segmentName', 'Unknown')
        display_name = get_segment_display_name(seg, level)
        seg_code = seg.get('segmentCode') or seg.get('segmentId')

        # ========== 收集数据 ==========
        revenue_values = []
        revenue_yoy = []
        ebita_values = []
        ebita_yoy = []
        margin_values = []  # EBITA利润率

        for period_idx, period in enumerate(periods):
            # 收入数据
            rev_metric = find_metric(seg, 'REVENUE', period)
            if rev_metric:
                revenue_values.append(rev_metric.get('value'))
                revenue_yoy.append(rev_metric.get('yoyGrowth'))
            else:
                revenue_values.append(None)
                revenue_yoy.append(None)

            # EBITA数据
            ebit_metric = find_metric(seg, 'ADJUSTED_EBITA', period) or find_metric(seg, 'OPERATING_INCOME', period)
            if ebit_metric:
                ebita_values.append(ebit_metric.get('value'))
                ebita_yoy.append(ebit_metric.get('yoyGrowth'))
                # 计算EBITA利润率 = EBITA / 收入 * 100
                revenue_val = revenue_values[period_idx]
                if revenue_val and revenue_val != 0 and ebit_metric.get('value'):
                    margin = (ebit_metric.get('value') / revenue_val) * 100
                else:
                    margin = None
                margin_values.append(margin)
            else:
                ebita_values.append(None)
                ebita_yoy.append(None)
                margin_values.append(None)

        # ========== 生成行 ==========
        # 1. 收入行
        if has_any_data(revenue_values):
            rows.append({
                'type': 'revenue',
                'level': level,
                'name': f"{display_name} - 收入",
                'values': revenue_values,
                'yoy_values': []  # YoY单独成行
            })

            # 2. 收入YoY行（如果有任何YoY数据）
            if has_any_data(revenue_yoy):
                rows.append({
                    'type': 'revenue_yoy',
                    'level': level,
                    'name': f"{display_name} - 收入YoY(%)",
                    'values': revenue_yoy,
                    'yoy_values': revenue_yoy
                })

        # 3. EBITA行
        if has_any_data(ebita_values):
            rows.append({
                'type': 'ebita',
                'level': level,
                'name': f"{display_name} - EBITA",
                'values': ebita_values,
                'yoy_values': []
            })

            # 4. EBITA YoY行
            if has_any_data(ebita_yoy):
                rows.append({
                    'type': 'ebita_yoy',
                    'level': level,
                    'name': f"{display_name} - EBITA YoY(%)",
                    'values': ebita_yoy,
                    'yoy_values': ebita_yoy
                })

            # 5. EBITA利润率行
            if has_any_data(margin_values):
                rows.append({
                    'type': 'ebita_margin',
                    'level': level,
                    'name': f"{display_name} - EBITA利润率(%)",
                    'values': margin_values,
                    'yoy_values': []
                })

    if logger:
        logger.info(f"过滤空数据行后，共生成 {len(rows)} 行数据")

    return rows


def find_metric(segment, metric_code, period):
    """在分部中查找指定指标和周期的数据"""
    for metric in segment.get('metrics', []):
        if metric.get('metricCode', '').upper() == metric_code.upper():
            if period is None or metric.get('period') == period:
                return metric
    return None


def format_value(val):
    """格式化数值 - 直接展示原始数据，不做单位转换"""
    if val is None:
        return '-'
    return val


def generate_excel_with_styling(segments, periods, output_file, ticker, logger=None):
    """生成带样式的Excel，体现层级关系"""
    if not HAS_OPENPYXL:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    wb = Workbook(write_only=False)
    ws = wb.active
    ws.title = "分部财务数据"

    # 基础样式
    header_font = Font(bold=True, name="微软雅黑", color=COLORS['header_font'])
    normal_font = Font(name="微软雅黑")
    level1_font = Font(bold=True, name="微软雅黑", color=COLORS['level1_font'])
    level2_font = Font(bold=True, name="微软雅黑", color=COLORS['level2_font'])
    level3_font = Font(name="微软雅黑", color=COLORS['level3_font'])
    red_font = Font(name="微软雅黑", color=COLORS['yoy_decline'], bold=True)
    green_font = Font(name="微软雅黑", color=COLORS['yoy_growth'], bold=True)

    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 构建数据行
    data_rows = build_data_rows(segments, periods, logger)

    # 写入表头
    # 结构：业务分部 | 指标 | [各期数值]
    headers = ['业务分部', '指标'] + periods
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = PatternFill(start_color=COLORS['header'],
                                end_color=COLORS['header'],
                                fill_type="solid")
        cell.alignment = center_align
        cell.border = thin_border

    # 写入数据
    row_idx = 2
    for data_row in data_rows:
        level = data_row['level']
        row_type = data_row['type']
        name = data_row['name']
        values = data_row['values']

        # 拆分名称：分部名称 和 指标名称（如 "L1 - 集团总览 - 收入" → "L1 - 集团总览", "收入"）
        name_parts = name.rsplit(' - ', 1)
        if len(name_parts) == 2:
            segment_name = name_parts[0]
            metric_name = name_parts[1]
        else:
            segment_name = name
            metric_name = ''

        # ========== 分部名称列（根据层级设置样式） ==========
        name_cell = ws.cell(row=row_idx, column=1)
        name_cell.value = segment_name
        name_cell.border = thin_border

        # 根据层级设置字体和背景色
        if level == 1:
            name_cell.font = level1_font
            name_cell.fill = PatternFill(start_color=COLORS['level1'],
                                        end_color=COLORS['level1'],
                                        fill_type="solid")
        elif level == 2:
            name_cell.font = level2_font
            name_cell.fill = PatternFill(start_color=COLORS['level2'],
                                        end_color=COLORS['level2'],
                                        fill_type="solid")
        else:
            name_cell.font = level3_font
            name_cell.fill = PatternFill(start_color=COLORS['level3'],
                                        end_color=COLORS['level3'],
                                        fill_type="solid")
        name_cell.alignment = left_align

        # ========== 指标类型列 ==========
        metric_cell = ws.cell(row=row_idx, column=2)
        metric_cell.value = metric_name
        metric_cell.border = thin_border
        metric_cell.alignment = center_align

        # ========== 数据列 ==========
        for col_offset, val in enumerate(values, 0):
            val_cell = ws.cell(row=row_idx, column=3 + col_offset)
            val_cell.border = thin_border
            val_cell.alignment = center_align

            # 设置数值和格式
            if val is not None and isinstance(val, (int, float)):
                val_cell.value = round(val, 2) if row_type.endswith('_yoy') or row_type == 'ebita_margin' else format_value(val)

                # YoY行的颜色高亮
                if row_type.endswith('_yoy'):
                    if val <= YOY_DECLINE_THRESHOLD:
                        val_cell.font = red_font
                    elif val >= YOY_GROWTH_THRESHOLD:
                        val_cell.font = green_font
                    else:
                        val_cell.font = normal_font
                else:
                    val_cell.font = normal_font

                # 数值格式
                if row_type.endswith('_yoy') or row_type == 'ebita_margin':
                    # 百分比类保留2位小数
                    val_cell.number_format = '0.00'
                elif val == int(val):
                    # 整数不带小数
                    val_cell.number_format = '#,##0'
                else:
                    # 小数保留2位
                    val_cell.number_format = '#,##0.00'
            else:
                val_cell.value = '-'
                val_cell.font = normal_font

            # 背景色根据行类型
            if row_type == 'revenue':
                val_cell.fill = PatternFill(start_color=COLORS['revenue'],
                                           end_color=COLORS['revenue'],
                                           fill_type="solid")
            elif row_type == 'ebita':
                val_cell.fill = PatternFill(start_color=COLORS['ebita'],
                                           end_color=COLORS['ebita'],
                                           fill_type="solid")
            elif row_type == 'ebita_margin':
                # 利润率行背景色（浅黄色）
                val_cell.fill = PatternFill(start_color=COLORS.get('ratio', 'FFF2CC'),
                                           end_color=COLORS.get('ratio', 'FFF2CC'),
                                           fill_type="solid")
            # YoY行背景色继承层级颜色

        # 移动到下一行
        row_idx += 1

    # 保存工作簿
    wb.save(output_file)

    if logger:
        logger.info(f"Excel saved: {output_file}")

    return output_file


def main():
    parser = argparse.ArgumentParser(description='生成分部业务财务报表Excel')
    parser.add_argument('ticker', nargs='?', help='股票代码，如 BABA, AAPL')
    parser.add_argument('--json', '-j', required=True, help='包含Segment数据的JSON文件路径（由Java Tool传入）')
    parser.add_argument('--output', '-o', help='输出Excel文件路径')
    parser.add_argument('--workspace', '-w', help='项目工作空间路径')
    args = parser.parse_args()

    # 确定项目根目录
    if args.workspace:
        # workspace已经是workspace目录了，不需要再加一层workspace
        project_root = args.workspace
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.abspath(os.path.join(script_dir, '..', '..'))

    # 确保日志和输出目录存在
    logs_dir = os.path.join(project_root, 'excels', 'logs')
    excels_dir = os.path.join(project_root, 'excels')
    for d in [logs_dir, excels_dir]:
        if not os.path.exists(d):
            os.makedirs(d)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    ticker = args.ticker or 'UNKNOWN'

    log_file = os.path.join(logs_dir, f"segment_{ticker}_{timestamp}.log")
    logger = setup_logging(log_file)

    log_stdout(f">>> Processing {ticker} 分部财务数据")

    try:
        # 从JSON文件加载数据（由Java Tool传入）
        if args.json and os.path.exists(args.json):
            with open(args.json, 'r', encoding='utf-8') as f:
                segments = json.load(f)
            logger.info(f"从JSON文件加载数据: {args.json}")
        else:
            log_stdout(f"ERROR: JSON文件不存在: {args.json}")
            return 1

        if not segments:
            log_stdout("ERROR: 分部数据为空")
            return 1

        logger.info(f"获取到 {len(segments)} 个顶层分部")

        # 获取所有周期
        periods = get_all_periods(segments)
        logger.info(f"包含 {len(periods)} 个财报周期: {periods}")

        if not periods:
            periods = ['Latest']

        # 确定输出路径
        if args.output:
            output_file = args.output
        else:
            output_file = os.path.join(excels_dir, f"{ticker}_segments_{timestamp}.xlsx")

        output_file = os.path.abspath(output_file)

        # 生成Excel
        generate_excel_with_styling(segments, periods, output_file, ticker, logger)

        # 最终输出
        result = {
            "status": "ok",
            "ticker": ticker,
            "segments_count": len(segments),
            "periods": periods,
            "excel_path": output_file,
            "log_path": os.path.abspath(log_file),
        }

        log_stdout("=" * 60)
        log_stdout(json.dumps(result, ensure_ascii=False))
        log_stdout("=" * 60)

        return 0

    except Exception as e:
        logger.exception("生成Excel失败")
        log_stdout(f"ERROR: {str(e)}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
